"""
admin/app.py

Flask ilovasi — blueprint yo'q, lekin route larni mantiqiy
bo'limlarga sarlavha bilan ajratilgan:

  ── Auth          /login  /logout
  ── Dashboard     /
  ── Users         /users  /users/<id>  (+ API)
  ── Questions     /questions  (+ API import/delete/export)
  ── Tests         /tests  /tests/<id>
  ── Leaderboard   /leaderboard
  ── Stats API     /api/stats/*
  ── Export        /export/*
  ── Extra routes  routes_extra.py dan ulangan

BUGFIX (v2.1):
  - export_questions: itertools import function ichidan module darajasiga ko'chirildi
  - flash alert kategoriyalar: 'warning' qo'shildi
"""

import io
import os
import sys
from datetime import datetime, timedelta
from itertools import groupby  # BUGFIX: modul darajasida import
from operator import attrgetter  # BUGFIX: modul darajasida import

from flask import (
    Flask, flash, jsonify, redirect,
    render_template, request, send_file, url_for,
)
from flask_login import (
    LoginManager, UserMixin,
    current_user, login_required, login_user, logout_user,
)
from sqlalchemy import desc, func
from werkzeug.security import check_password_hash, generate_password_hash

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config

# ── App setup ────────────────────────────────────────────────────────────────

app = Flask(__name__)
app.config["SECRET_KEY"]           = config.SECRET_KEY
app.config["DEBUG"]                = config.FLASK_DEBUG
app.config["MAX_CONTENT_LENGTH"]   = 16 * 1024 * 1024

# ── Auth setup ───────────────────────────────────────────────────────────────

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view    = "login"
login_manager.login_message = "Iltimos, tizimga kiring."

ADMIN_CREDENTIALS = {
    os.getenv("ADMIN_USERNAME", "admin"): generate_password_hash(
        os.getenv("ADMIN_PASSWORD", "dtm_admin_2025")
    )
}


class AdminUser(UserMixin):
    def __init__(self, username: str):
        self.id       = username
        self.username = username


@login_manager.user_loader
def load_user(user_id):
    if user_id in ADMIN_CREDENTIALS:
        return AdminUser(user_id)
    return None


def get_db():
    from database.db import Session
    return Session()


# ══════════════════════════════════════════════════════════════════════════════
# AUTH
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))

    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        if username in ADMIN_CREDENTIALS and \
                check_password_hash(ADMIN_CREDENTIALS[username], password):
            login_user(AdminUser(username), remember=True)
            return redirect(url_for("dashboard"))
        error = "Login yoki parol noto'g'ri!"

    return render_template("login.html", error=error)


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/")
@login_required
def dashboard():
    db = get_db()
    try:
        from database.models import Question, Score, User, UserTestParticipation

        total_users      = db.query(func.count(User.id)).scalar() or 0
        total_tests_done = db.query(func.count(Score.id)).scalar() or 0
        active_tests     = db.query(func.count(UserTestParticipation.id)).filter(
            UserTestParticipation.status == "active"
        ).scalar() or 0
        total_questions  = db.query(func.count(Question.id)).scalar() or 0

        week_ago       = datetime.utcnow() - timedelta(days=7)
        new_users_week = db.query(func.count(User.id)).filter(
            User.created_at >= week_ago
        ).scalar() or 0

        today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        tests_today = db.query(func.count(Score.id)).filter(
            Score.created_at >= today_start
        ).scalar() or 0

        avg_score_raw = db.query(func.avg(Score.score)).scalar()
        avg_score     = round(float(avg_score_raw), 1) if avg_score_raw else 0

        recent_users = db.query(User).order_by(desc(User.created_at)).limit(5).all()
        top_scores   = db.query(Score).order_by(desc(Score.score)).limit(5).all()

        # Kunlik registratsiya grafigi (7 kun)
        daily_reg = []
        for i in range(6, -1, -1):
            day       = datetime.utcnow() - timedelta(days=i)
            day_start = day.replace(hour=0,  minute=0,  second=0,  microsecond=0)
            day_end   = day.replace(hour=23, minute=59, second=59, microsecond=999999)
            count     = db.query(func.count(User.id)).filter(
                User.created_at.between(day_start, day_end)
            ).scalar() or 0
            daily_reg.append({"date": day.strftime("%d.%m"), "count": count})

        return render_template(
            "dashboard.html",
            total_users=total_users,
            total_tests_done=total_tests_done,
            active_tests=active_tests,
            total_questions=total_questions,
            new_users_week=new_users_week,
            tests_today=tests_today,
            avg_score=avg_score,
            recent_users=recent_users,
            top_scores=top_scores,
            daily_reg=daily_reg,
        )
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════════════════════
# USERS
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/users")
@login_required
def users():
    db = get_db()
    try:
        from database.models import Region, Score, User

        page          = request.args.get("page", 1, type=int)
        search        = request.args.get("search", "").strip()
        region_filter = request.args.get("region", "", type=str)
        per_page      = 20

        query = db.query(User)
        if search:
            query = query.filter(
                User.first_name.ilike(f"%{search}%") |
                User.last_name.ilike(f"%{search}%")  |
                User.phone.ilike(f"%{search}%")
            )
        if region_filter:
            query = query.filter(User.region_id == region_filter)

        total     = query.count()
        user_list = (
            query.order_by(desc(User.created_at))
            .offset((page - 1) * per_page)
            .limit(per_page)
            .all()
        )

        # Bitta GROUP BY so'rov — 40 ta alohida so'rov o'rniga
        user_ids = [u.id for u in user_list]
        stats_rows = (
            db.query(
                Score.user_id,
                func.count(Score.id).label("test_count"),
                func.max(Score.score).label("best_score"),
            )
            .filter(Score.user_id.in_(user_ids))
            .group_by(Score.user_id)
            .all()
        )
        user_stats = {
            row.user_id: {
                "test_count": row.test_count,
                "best_score": round(float(row.best_score), 1) if row.best_score else 0,
            }
            for row in stats_rows
        }
        for u in user_list:
            if u.id not in user_stats:
                user_stats[u.id] = {"test_count": 0, "best_score": 0}

        regions     = db.query(Region).all()
        total_pages = (total + per_page - 1) // per_page

        return render_template(
            "users.html",
            users=user_list,
            user_stats=user_stats,
            total=total,
            page=page,
            total_pages=total_pages,
            search=search,
            region_filter=region_filter,
            regions=regions,
        )
    finally:
        db.close()


@app.route("/users/<int:user_id>")
@login_required
def user_detail(user_id):
    db = get_db()
    try:
        from database.models import Score, User, UserTestParticipation
        from sqlalchemy.orm import joinedload

        user = db.query(User).options(
            joinedload(User.region),
            joinedload(User.district),
            joinedload(User.direction),
        ).filter(User.id == user_id).first()

        if not user:
            flash("Foydalanuvchi topilmadi", "error")
            return redirect(url_for("users"))

        scores = (
            db.query(Score)
            .filter(Score.user_id == user_id)
            .order_by(desc(Score.created_at))
            .limit(10)
            .all()
        )
        participations = (
            db.query(UserTestParticipation)
            .filter(UserTestParticipation.user_id == user_id)
            .order_by(desc(UserTestParticipation.joined_at))
            .limit(10)
            .all()
        )

        return render_template(
            "user_detail.html",
            user=user,
            scores=scores,
            participations=participations,
        )
    finally:
        db.close()


@app.route("/api/users/<int:user_id>/delete", methods=["POST"])
@login_required
def delete_user(user_id):
    db = get_db()
    try:
        from database.models import (
            Leaderboard, ReferralInvite, ReferralLink,
            Score, User, UserAnswer, UserTestParticipation,
        )

        # 1. Participation va unga bog'liq ma'lumotlar
        participations    = db.query(UserTestParticipation).filter(
            UserTestParticipation.user_id == user_id
        ).all()
        participation_ids = [p.id for p in participations]

        if participation_ids:
            db.query(Score).filter(
                Score.participation_id.in_(participation_ids)
            ).delete(synchronize_session=False)
            db.query(UserAnswer).filter(
                UserAnswer.participation_id.in_(participation_ids)
            ).delete(synchronize_session=False)

        # 2. To'g'ridan-to'g'ri user ga bog'liq ma'lumotlar
        db.query(Score).filter(Score.user_id == user_id).delete(synchronize_session=False)
        db.query(Leaderboard).filter(Leaderboard.user_id == user_id).delete(synchronize_session=False)
        db.query(UserAnswer).filter(UserAnswer.user_id == user_id).delete(synchronize_session=False)
        db.query(UserTestParticipation).filter(
            UserTestParticipation.user_id == user_id
        ).delete(synchronize_session=False)

        # 3. Referal ma'lumotlar (avval invites, keyin link — FK tartib)
        link = db.query(ReferralLink).filter(ReferralLink.user_id == user_id).first()
        if link:
            db.query(ReferralInvite).filter(
                ReferralInvite.referral_link_id == link.id
            ).delete(synchronize_session=False)
            db.query(ReferralLink).filter(ReferralLink.id == link.id).delete(synchronize_session=False)

        # Bu user boshqa havola orqali kelganmi — invite yozuvini ham tozala
        db.query(ReferralInvite).filter(
            ReferralInvite.invited_user_id == user_id
        ).delete(synchronize_session=False)

        # 4. Foydalanuvchini o'chirish
        db.query(User).filter(User.id == user_id).delete(synchronize_session=False)

        db.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.rollback()
        return jsonify({"success": False, "error": str(e)})
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════════════════════
# QUESTIONS
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/questions")
@login_required
def questions():
    db = get_db()
    try:
        from database.models import Question, Subject

        page           = request.args.get("page", 1, type=int)
        subject_filter = request.args.get("subject", 0, type=int)
        search         = request.args.get("search", "").strip()
        per_page       = 25

        query = db.query(Question)
        if subject_filter:
            query = query.filter(Question.subject_id == subject_filter)
        if search:
            query = query.filter(Question.text_uz.ilike(f"%{search}%"))

        total         = query.count()
        question_list = (
            query.order_by(Question.subject_id, Question.id)
            .offset((page - 1) * per_page)
            .limit(per_page)
            .all()
        )
        subjects       = db.query(Subject).all()
        subject_counts = {
            s.id: db.query(func.count(Question.id)).filter(
                Question.subject_id == s.id
            ).scalar() or 0
            for s in subjects
        }

        return render_template(
            "questions.html",
            questions=question_list,
            subjects=subjects,
            subject_counts=subject_counts,
            subject_filter=subject_filter,
            search=search,
            total=total,
            page=page,
            total_pages=(total + per_page - 1) // per_page,
        )
    finally:
        db.close()


@app.route("/api/questions/import", methods=["POST"])
@login_required
def import_questions():
    if "file" not in request.files:
        return jsonify({"success": False, "error": "Fayl yuklanmadi"})

    file = request.files["file"]
    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"success": False, "error": "Faqat Excel fayl"})

    subject_id = request.form.get("subject_id", type=int)
    if not subject_id:
        return jsonify({"success": False, "error": "Fan tanlanmadi"})

    try:
        import openpyxl
        from database.models import Question, Subject

        wb  = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws  = wb.active
        db  = get_db()
        added, errors = 0, []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            try:
                text    = str(row[0]).strip() if row[0] else ""
                opt_a   = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                opt_b   = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                opt_c   = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                opt_d   = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                correct = str(row[5]).strip().upper() if len(row) > 5 and row[5] else ""

                if not all([text, opt_a, opt_b, opt_c, opt_d]) or correct not in "ABCD":
                    errors.append(f"Qator {row_idx}: to'liq emas")
                    continue

                db.add(Question(
                    subject_id=subject_id,
                    text_uz=text, text_oz=text, text_ru=text,
                    option_a=opt_a, option_b=opt_b,
                    option_c=opt_c, option_d=opt_d,
                    correct_answer=correct,
                ))
                added += 1
            except Exception as e:
                errors.append(f"Qator {row_idx}: {e}")

        db.commit()

        subj = db.query(Subject).filter(Subject.id == subject_id).first()
        if subj:
            subj.question_count = db.query(func.count(Question.id)).filter(
                Question.subject_id == subject_id
            ).scalar()
            db.commit()
        db.close()

        return jsonify({
            "success": True,
            "added":   added,
            "errors":  errors[:10],
            "message": f"{added} ta savol qo'shildi",
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/questions/<int:q_id>/delete", methods=["POST"])
@login_required
def delete_question(q_id):
    db = get_db()
    try:
        from database.models import Question, UserAnswer
        db.query(UserAnswer).filter(UserAnswer.question_id == q_id).delete()
        db.query(Question).filter(Question.id == q_id).delete()
        db.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.rollback()
        return jsonify({"success": False, "error": str(e)})
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════════════════════
# TESTS
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/tests")
@login_required
def tests():
    db = get_db()
    try:
        from database.models import Score, TestSession, UserTestParticipation

        test_sessions = (
            db.query(TestSession)
            .order_by(desc(TestSession.created_at))
            .limit(20)
            .all()
        )

        session_stats = {}
        for ts in test_sessions:
            participants = db.query(func.count(UserTestParticipation.id)).filter(
                UserTestParticipation.test_session_id == ts.id
            ).scalar() or 0

            completed = db.query(func.count(UserTestParticipation.id)).filter(
                UserTestParticipation.test_session_id == ts.id,
                UserTestParticipation.status == "completed",
            ).scalar() or 0

            avg = (
                db.query(func.avg(Score.score))
                .join(UserTestParticipation,
                      Score.participation_id == UserTestParticipation.id)
                .filter(UserTestParticipation.test_session_id == ts.id)
                .scalar()
            )
            if avg is None:
                avg = (
                    db.query(func.avg(Score.score))
                    .join(UserTestParticipation,
                          UserTestParticipation.user_id == Score.user_id)
                    .filter(UserTestParticipation.test_session_id == ts.id)
                    .scalar()
                )

            session_stats[ts.id] = {
                "participants": participants,
                "completed":    completed,
                "avg_score":    round(float(avg), 1) if avg else 0,
            }

        return render_template(
            "tests.html",
            test_sessions=test_sessions,
            session_stats=session_stats,
        )
    finally:
        db.close()


@app.route("/tests/<int:session_id>")
@login_required
def test_detail(session_id):
    db = get_db()
    try:
        from database.models import Score, TestSession, UserTestParticipation
        from sqlalchemy.orm import joinedload

        session = db.query(TestSession).filter(TestSession.id == session_id).first()
        if not session:
            flash("Session topilmadi", "error")
            return redirect(url_for("tests"))

        participations = (
            db.query(UserTestParticipation)
            .options(joinedload(UserTestParticipation.user))
            .filter(UserTestParticipation.test_session_id == session_id)
            .order_by(UserTestParticipation.joined_at)
            .all()
        )

        scores_map = {}
        for p in participations:
            score = db.query(Score).filter(Score.participation_id == p.id).first()
            if not score:
                score = (
                    db.query(Score)
                    .filter(Score.user_id == p.user_id)
                    .order_by(desc(Score.created_at))
                    .first()
                )
            scores_map[p.user_id] = score

        return render_template(
            "test_detail.html",
            session=session,
            participations=participations,
            scores_map=scores_map,
        )
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════════════════════
# LEADERBOARD
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/leaderboard")
@login_required
def leaderboard():
    db = get_db()
    try:
        from database.models import Direction, Leaderboard, User
        from sqlalchemy.orm import joinedload

        page             = request.args.get("page", 1, type=int)
        direction_filter = request.args.get("direction", "").strip()
        period_filter    = request.args.get("period", "all_time")
        per_page         = 25

        query = (
            db.query(Leaderboard)
            .options(
                joinedload(Leaderboard.user),
                joinedload(Leaderboard.direction),
            )
            .filter(Leaderboard.period == period_filter)
        )
        if direction_filter:
            query = query.filter(Leaderboard.direction_id == direction_filter)

        total   = query.count()
        entries = (
            query.order_by(Leaderboard.rank)
            .offset((page - 1) * per_page)
            .limit(per_page)
            .all()
        )

        directions  = db.query(Direction).order_by(Direction.name_uz).all()
        total_pages = (total + per_page - 1) // per_page

        period_names = {
            "daily":    "Kunlik",
            "weekly":   "Haftalik",
            "all_time": "Barcha vaqt",
        }

        return render_template(
            "leaderboard.html",
            entries=entries,
            directions=directions,
            direction_filter=direction_filter,
            period_filter=period_filter,
            period_names=period_names,
            total=total,
            page=page,
            total_pages=total_pages,
        )
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════════════════════
# STATS API
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/stats/daily")
@login_required
def stats_daily():
    db = get_db()
    try:
        from database.models import Score, User

        days   = request.args.get("days", 7, type=int)
        result = {"registrations": [], "tests": []}

        for i in range(days - 1, -1, -1):
            day   = datetime.utcnow() - timedelta(days=i)
            ds    = day.replace(hour=0,  minute=0,  second=0,  microsecond=0)
            de    = day.replace(hour=23, minute=59, second=59, microsecond=999999)
            label = day.strftime("%d.%m")

            reg_count  = db.query(func.count(User.id)).filter(
                User.created_at.between(ds, de)
            ).scalar() or 0
            test_count = db.query(func.count(Score.id)).filter(
                Score.created_at.between(ds, de)
            ).scalar() or 0

            result["registrations"].append({"date": label, "count": reg_count})
            result["tests"].append({"date": label, "count": test_count})

        return jsonify(result)
    finally:
        db.close()


@app.route("/api/stats/subjects")
@login_required
def stats_subjects():
    db = get_db()
    try:
        from database.models import Question, Subject

        subjects = db.query(Subject).all()
        return jsonify([
            {
                "name":  s.name_uz,
                "count": db.query(func.count(Question.id)).filter(
                    Question.subject_id == s.id
                ).scalar() or 0,
            }
            for s in subjects
        ])
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/export/users")
@login_required
def export_users():
    try:
        import openpyxl
        from database.models import Score, User

        db        = get_db()
        users_all = db.query(User).order_by(User.id).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Foydalanuvchilar"
        ws.append([
            "ID", "Ism", "Familiya", "Telefon", "Viloyat", "Tuman",
            "Yo'nalish", "Test soni", "Eng yaxshi ball", "Ro'yxat sanasi",
        ])

        for u in users_all:
            scores = db.query(Score).filter(Score.user_id == u.id).all()
            best   = max((s.score for s in scores), default=0)
            ws.append([
                u.id, u.first_name, u.last_name or "", u.phone,
                u.region.name_uz    if u.region    else "",
                u.district.name_uz  if u.district  else "",
                u.direction.name_uz if u.direction  else "",
                len(scores), best,
                u.created_at.strftime("%d.%m.%Y %H:%M"),
            ])
        db.close()

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"users_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        flash(f"Export xatosi: {e}", "error")
        return redirect(url_for("users"))


@app.route("/export/scores")
@login_required
def export_scores():
    try:
        import openpyxl
        from database.models import Score

        db     = get_db()
        scores = db.query(Score).order_by(desc(Score.score)).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Natijalar"
        ws.append([
            "#", "Ism", "Familiya", "Telefon", "Yo'nalish",
            "Ball", "To'g'ri", "Jami", "Foiz", "Sana",
        ])

        for i, s in enumerate(scores, 1):
            u   = s.user
            pct = round(s.correct_count / s.total_questions * 100, 1) if s.total_questions else 0
            ws.append([
                i,
                u.first_name if u else "",
                u.last_name  if u else "",
                u.phone      if u else "",
                u.direction.name_uz if u and u.direction else "",
                s.score, s.correct_count, s.total_questions, pct,
                s.created_at.strftime("%d.%m.%Y %H:%M"),
            ])
        db.close()

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"scores_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        flash(f"Export xatosi: {e}", "error")
        return redirect(url_for("leaderboard"))


@app.route("/export/questions")
@login_required
def export_questions():
    """
    BUGFIX: itertools va openpyxl.styles importlar modul darajasiga ko'chirildi.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from database.models import Question, Subject

        subject_filter = request.args.get("subject", 0, type=int)

        db             = get_db()
        query = db.query(Question)
        if subject_filter:
            query = query.filter(Question.subject_id == subject_filter)
        questions_all = query.order_by(Question.subject_id, Question.id).all()
        subjects_map  = {s.id: s.name_uz for s in db.query(Subject).all()}
        db.close()

        # ── Stil konstantalar ───────────────────────────────────────────────
        HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        HEADER_FILL  = PatternFill("solid", start_color="1E2A78")
        CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
        LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        THIN         = Side(style="thin", color="BDBDBD")
        BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        SUBJECT_COLORS = {
            1: "E3F2FD", 2: "FFF8E1", 3: "F3E5F5", 4: "E8F5E9", 5: "FFFDE7",
            6: "FCE4EC", 7: "E0F2F1", 8: "E1F5FE", 9: "F9FBE7", 10: "FBE9E7",
        }
        SUBJECT_HDR_COLORS = {
            1: "1565C0", 2: "F57F17", 3: "6A1B9A", 4: "1B5E20", 5: "BF360C",
            6: "880E4F", 7: "004D40", 8: "01579B", 9: "33691E", 10: "B71C1C",
        }
        IMPORT_HEADERS = [
            "Savol matni", "Variant A", "Variant B", "Variant C", "Variant D",
            "To'g'ri javob (A/B/C/D)",
        ]
        COL_WIDTHS = [65, 32, 32, 32, 32, 24]

        # ── "Barcha savollar" sheet ─────────────────────────────────────────
        wb     = openpyxl.Workbook()
        ws_all = wb.active
        ws_all.title = "Barcha savollar"

        for col, (h, w) in enumerate(zip(IMPORT_HEADERS + ["Fan nomi"],
                                          COL_WIDTHS + [18]), 1):
            cell = ws_all.cell(row=1, column=col, value=h)
            cell.font, cell.fill = HEADER_FONT, HEADER_FILL
            cell.alignment, cell.border = CENTER_ALIGN, BORDER
            ws_all.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
        ws_all.row_dimensions[1].height = 28
        ws_all.freeze_panes = "A2"

        for r, q in enumerate(questions_all, 2):
            row_fill = PatternFill("solid", start_color=SUBJECT_COLORS.get(q.subject_id, "FFFFFF"))
            vals = [
                q.text_uz, q.option_a, q.option_b, q.option_c, q.option_d,
                q.correct_answer, subjects_map.get(q.subject_id, str(q.subject_id)),
            ]
            for col, val in enumerate(vals, 1):
                cell            = ws_all.cell(row=r, column=col, value=val)
                cell.fill       = row_fill
                cell.border     = BORDER
                cell.alignment  = CENTER_ALIGN if col in (6, 7) else LEFT_ALIGN
                if col == 6:
                    cell.font = Font(
                        name="Arial", bold=True,
                        color=SUBJECT_HDR_COLORS.get(q.subject_id, "1B5E20"), size=11,
                    )
            ws_all.row_dimensions[r].height = 42

        note_row = len(questions_all) + 3
        ws_all.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=7)
        note = ws_all.cell(row=note_row, column=1,
                           value="Importda faqat A-F ustunlarini ishlating. 1-qator sarlavha.")
        note.font      = Font(name="Arial", italic=True, color="616161", size=9)
        note.alignment = LEFT_ALIGN

        # ── Har fan uchun alohida sheet ─────────────────────────────────────
        # BUGFIX: groupby va attrgetter modul darajasida import qilingan
        for subject_id, grp in groupby(
            sorted(questions_all, key=attrgetter("subject_id")),
            key=attrgetter("subject_id"),
        ):
            grp_list  = list(grp)
            subj_name = subjects_map.get(subject_id, f"Fan-{subject_id}")
            hdr_color = SUBJECT_HDR_COLORS.get(subject_id, "1E2A78")
            row_color = SUBJECT_COLORS.get(subject_id, "FFFFFF")
            ws        = wb.create_sheet(title=subj_name[:28])

            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
            banner = ws.cell(row=1, column=1,
                             value=f"📚 {subj_name}  —  {len(grp_list)} ta savol")
            banner.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
            banner.fill      = PatternFill("solid", start_color=hdr_color)
            banner.alignment = Alignment(horizontal="center", vertical="center")
            banner.border    = BORDER
            ws.row_dimensions[1].height = 32

            for col, (h, w) in enumerate(zip(IMPORT_HEADERS, COL_WIDTHS), 1):
                cell = ws.cell(row=2, column=col, value=h)
                cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                cell.fill      = PatternFill("solid", start_color=hdr_color)
                cell.alignment = CENTER_ALIGN
                cell.border    = BORDER
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
            ws.row_dimensions[2].height = 26
            ws.freeze_panes = "A3"

            row_fill = PatternFill("solid", start_color=row_color)
            for r, q in enumerate(grp_list, 3):
                for col, val in enumerate(
                    [q.text_uz, q.option_a, q.option_b, q.option_c, q.option_d, q.correct_answer], 1
                ):
                    cell           = ws.cell(row=r, column=col, value=val)
                    cell.fill      = row_fill
                    cell.border    = BORDER
                    cell.alignment = CENTER_ALIGN if col == 6 else LEFT_ALIGN
                    if col == 6:
                        cell.font = Font(name="Arial", bold=True, color=hdr_color, size=11)
                ws.row_dimensions[r].height = 42

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        suffix   = f"_fan{subject_filter}" if subject_filter else "_barchasi"
        filename = f"savollar{suffix}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        flash(f"Export xatosi: {e}", "error")
        return redirect(url_for("questions"))


# ══════════════════════════════════════════════════════════════════════════════
# EXTRA ROUTES
# ══════════════════════════════════════════════════════════════════════════════

from admin.routes_extra import register_extra_routes
register_extra_routes(app)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=config.FLASK_DEBUG)