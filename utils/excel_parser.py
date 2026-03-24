"""
utils/excel_parser.py — Fanlar majmuasi Excel faylini o'qish

Fayl qidiriladigan joylar (tartib bo'yicha):
  1. data/Fanlar_majmuasi_2025-2026.xlsx  ← tavsiya etilgan joy
  2. Fanlar_majmuasi_2025-2026.xlsx        ← loyiha ildizida
"""
import os
import re
from typing import List, Dict

try:
    import openpyxl
except ImportError:
    raise ImportError("openpyxl o'rnatilmagan. pip install openpyxl")

# Loyiha ildizi (utils/ ning parenti)
ROOT_DIR = os.path.dirname(os.path.abspath(__file__))

EXCEL_FILES = [
    os.path.join(ROOT_DIR, "Fanlar_majmuasi_2025-2026.xlsx"),
]

SUBJECT_MAP = {
    "matematika": 1,
    "fizika": 2,
    "kimyo": 3,
    "biologiya": 4,
    "tarix": 5,
    "ona tili": 6,
    "ona tili va adabiyoti": 6,
    "o'zbek tili va adabiyoti": 6,
    "oʻzbek tili va adabiyoti": 6,
    "adabiyot": 7,
    "geografiya": 8,
    "ingliz tili": 9,
    "chet tili": 9,
    "nemis tili": 9,
    "fransuz tili": 9,
    "rus tili": 10,
    "rus tili va adabiyoti": 10,
    "qirgʻiz tili va adabiyoti": 9,
    "qozoq tili va adabiyoti": 9,
    "tojik tili va adabiyoti": 9,
    "turkman tili va adabiyoti": 9,
    "qoraqalpoq tili va adabiyoti": 9,
    "kasbiy (ijodiy imtihon)": 7,
    "kasbiy (ijodiy) imtihon": 7,
    "kasbiy": 7,
    "huquqshunoslik fanlari": 5,
    "huquqshunoslik": 5,
}


def get_subject_id(name: str) -> int:
    if not name:
        return 1
    clean = name.lower().strip()
    if clean in SUBJECT_MAP:
        return SUBJECT_MAP[clean]
    for key, sid in SUBJECT_MAP.items():
        if key in clean:
            return sid
    print(f"  [WARN] Noma'lum fan: '{name}' — Matematika (1) deb olindi")
    return 1


def _clean(val) -> str:
    return str(val).strip() if val is not None else ""


def _safe_get(row, index) -> str:
    if index is None or index < 0 or index >= len(row):
        return ""
    return _clean(row[index])


def _detect_columns(headers: List[str]) -> Dict[str, int]:
    mapping = {}
    for i, h in enumerate(headers):
        h_low = h.lower().strip()
        if not h_low:
            continue
        if any(k in h_low for k in ["kod", "code", "raqam", "shifr"]) and "code" not in mapping:
            mapping["code"] = i
        elif any(k in h_low for k in ["nom", "name", "yo'nalish", "yonalish", "mutaxassis"]) and "name" not in mapping:
            mapping["name"] = i
        elif any(k in h_low for k in ["1-fan", "fan 1", "subject1", "birinchi", "1 fan"]) and "subject1" not in mapping:
            mapping["subject1"] = i
        elif any(k in h_low for k in ["2-fan", "fan 2", "subject2", "ikkinchi", "2 fan"]) and "subject2" not in mapping:
            mapping["subject2"] = i
    return mapping


def _guess_columns(rows: List) -> Dict[str, int]:
    for row in rows[:10]:
        for i, cell in enumerate(row):
            val = re.sub(r'\s+', '', _clean(cell))
            if re.match(r'^\d{6,9}$', val):
                return {"code": i, "name": i + 1, "subject1": i + 2, "subject2": i + 3}
    return {}


def _parse_sheet(ws) -> List[Dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_row_idx = None
    col_map = {}
    for idx, row in enumerate(rows[:5]):
        headers = [_clean(c) for c in row]
        col_map = _detect_columns(headers)
        if len(col_map) >= 3:
            header_row_idx = idx
            print(f"  Sarlavha qatori: {idx + 1} | {col_map}")
            break

    if header_row_idx is None:
        col_map = _guess_columns(rows)
        header_row_idx = 0
        if not col_map:
            return []

    directions = []
    seen_codes = set()

    for row in rows[header_row_idx + 1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        code  = re.sub(r'\s+', '', _safe_get(row, col_map.get("code")))
        name  = _safe_get(row, col_map.get("name"))
        subj1 = _safe_get(row, col_map.get("subject1"))
        subj2 = _safe_get(row, col_map.get("subject2"))

        if not re.match(r'^\d{6,9}$', code) or not name or code in seen_codes:
            continue
        seen_codes.add(code)

        directions.append({
            "code":       code,
            "name":       name,
            "subject1":   subj1,
            "subject2":   subj2,
            "subject1_id": get_subject_id(subj1),
            "subject2_id": get_subject_id(subj2),
        })

    return directions


def parse_directions_from_excel() -> List[Dict]:
    for filepath in EXCEL_FILES:
        if not os.path.exists(filepath):
            continue

        print(f"\n📂 O'qilmoqda: {os.path.relpath(filepath, ROOT_DIR)}")
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            all_directions = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                dirs = _parse_sheet(ws)
                print(f"  Sheet '{sheet_name}': {len(dirs)} ta yo'nalish")
                all_directions.extend(dirs)
            wb.close()

            if all_directions:
                seen = set()
                unique = [d for d in all_directions
                          if d["code"] not in seen and not seen.add(d["code"])]
                print(f"  ✅ Jami: {len(unique)} ta unikal yo'nalish")
                return unique
        except Exception as e:
            print(f"  [ERROR] {e}")
            continue

    print(f"\n⚠️  Excel fayl topilmadi!")
    print(f"   Fayl joyi: data/Fanlar_majmuasi_2025-2026.xlsx")
    print(f"   Fallback 5 ta yo'nalish ishlatiladi.\n")
    return _fallback_directions()


def _fallback_directions() -> List[Dict]:
    return [
        {"code": "60610400", "name": "Dasturiy injiniring",    "subject1": "Matematika", "subject2": "Fizika",   "subject1_id": 1, "subject2_id": 2},
        {"code": "60610500", "name": "Sun'iy intellekt",        "subject1": "Matematika", "subject2": "Fizika",   "subject1_id": 1, "subject2_id": 2},
        {"code": "60540100", "name": "Matematika",              "subject1": "Matematika", "subject2": "Fizika",   "subject1_id": 1, "subject2_id": 2},
        {"code": "60110100", "name": "Pedagogika",              "subject1": "Tarix",      "subject2": "Ona tili", "subject1_id": 5, "subject2_id": 6},
        {"code": "60420100", "name": "Yurisprudensiya",         "subject1": "Tarix",      "subject2": "Chet tili","subject1_id": 5, "subject2_id": 9},
    ]